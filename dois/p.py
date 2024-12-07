from openpyxl import load_workbook
from openpyxl.styles import PatternFill

def load_worksheet(file_path):
    workbook = load_workbook(file_path)
    return workbook.active

def get_column_indices(worksheet):
    return {cell.value: idx for idx, cell in enumerate(worksheet[1])}

def update_cadastro(cadastro_ws, definicao_ws, col_indices_def, col_indices_cad):
    # Mapeia NCM ou variações para seus dados correspondentes
    definicao_ncm_map = {}
    definicao_variacao_map = {}
    
    for row in definicao_ws.iter_rows(min_row=2, values_only=True):
        ncm_or_variacao = str(row[col_indices_def['VARIACAO OU NCM']]).strip().lower()
        if ncm_or_variacao.replace('.', '').isdigit():  # Verifica se é NCM, considerando pontos
            definicao_ncm_map[ncm_or_variacao] = (
                row[col_indices_def['ALIQ ICMS']],
                row[col_indices_def['CST ICMS']],
                row[col_indices_def['CST PIS E COFINS']],
                row[col_indices_def['CBENEF']]
            )
        else:
            key_parts = ncm_or_variacao.split()
            definicao_variacao_map[tuple(key_parts)] = (
                row[col_indices_def['ALIQ ICMS']],
                row[col_indices_def['CST ICMS']],
                row[col_indices_def['CST PIS E COFINS']],
                row[col_indices_def['CBENEF']]
            )
    
    light_green_fill = PatternFill(start_color='90EE90', end_color='90EE90', fill_type='solid')

    for row in cadastro_ws.iter_rows(min_row=2):
        descricao_cell = row[col_indices_cad['DESCRICAO']].value
        descricao = str(descricao_cell).strip().lower() if descricao_cell else ""
        
        ncm = str(row[col_indices_cad['NCM']].value or "").strip().lower()
        
        # Primeiro, tenta encontrar o NCM
        best_data = definicao_ncm_map.get(ncm)
        
        # Se não encontrou pelo NCM, tenta pela descrição
        if not best_data:
            best_score = -1
            for key_parts, def_data in definicao_variacao_map.items():
                if all(part in descricao for part in key_parts):
                    match_score = sum(len(part) for part in key_parts)
                    if match_score > best_score:
                        best_score = match_score
                        best_data = def_data

        # Atualiza as células se os dados adequados forem encontrados
        if best_data:
            updated = False
            for idx, col in enumerate(['ALIQ ICMS', 'CST ICMS', 'CST PIS E COFINS', 'CBENEF']):
                cell = row[col_indices_cad[col]]
                if cell.value != best_data[idx]:
                    cell.value = best_data[idx]
                    cell.fill = light_green_fill
                    updated = True

            if updated:
                row[col_indices_cad['DESCRICAO']].fill = light_green_fill

def executar_codigo(caminho, progress_callback=None):
    cadastro_file = f"{caminho}\\CADASTRO.xlsx"
    definicao_file = f"{caminho}\\DEFINICAO.xlsx"

    cadastro_ws = load_worksheet(cadastro_file)
    definicao_ws = load_worksheet(definicao_file)

    definicao_col_indices = get_column_indices(definicao_ws)
    cadastro_col_indices = get_column_indices(cadastro_ws)

    update_cadastro(cadastro_ws, definicao_ws, definicao_col_indices, cadastro_col_indices)

    cadastro_ws.parent.save(cadastro_file)

    if progress_callback:
        progress_callback(100)